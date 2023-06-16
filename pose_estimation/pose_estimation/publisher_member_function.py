
# ============ IMPORT PACKAGES =================
# Packages for main code
import cv2 as cv
import numpy as np
import mediapipe as mp
import glob
import os
import yaml
import time
import openpyxl
import scipy

import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
from scipy.signal import filtfilt, butter

# Packages for ROS2
import rclpy
from rclpy.node import Node
   
from geometry_msgs.msg import Point
from d2w_interfaces.msg import SkeletonPoints


# MediaPipe
mp_drawing = mp.solutions.drawing_utils
mp_drawing_style = mp.solutions.drawing_styles
mp_pose = mp.solutions.pose


# ============ DEFINE CLASSES =================

class PosePublisher(Node):

    def __init__(self):
        super().__init__('pose_publisher')
        self.pub_skeleton = self.create_publisher(SkeletonPoints, '/d2w/patient/skeleton', 10)   
       
    def publish_skeleton(self, pose_array):
        """ This method publishes the whole body but the face landmarks
        """
        
        msg = SkeletonPoints()
        
        msg.r_shoulder_ref.position.x = float(pose_array[0][0])
        msg.r_shoulder_ref.position.z = float(pose_array[0][1])
        msg.r_shoulder_ref.position.y = float(pose_array[0][2])

        msg.l_shoulder_ref.position.x = float(pose_array[1][0])
        msg.l_shoulder_ref.position.z = float(pose_array[1][1])
        msg.l_shoulder_ref.position.y = float(pose_array[1][2])

        msg.r_elbow_ref.position.x = float(pose_array[2][0])
        msg.r_elbow_ref.position.z = float(pose_array[2][1])
        msg.r_elbow_ref.position.y = float(pose_array[2][2])

        msg.l_elbow_ref.position.x = float(pose_array[3][0])
        msg.l_elbow_ref.position.z = float(pose_array[3][1])
        msg.l_elbow_ref.position.y = float(pose_array[3][2])

        msg.r_wrist_ref.position.x = float(pose_array[4][0])
        msg.r_wrist_ref.position.z = float(pose_array[4][1])
        msg.r_wrist_ref.position.y = float(pose_array[4][2])

        msg.l_wrist_ref.position.x = float(pose_array[5][0])
        msg.l_wrist_ref.position.z = float(pose_array[5][1])
        msg.l_wrist_ref.position.y = float(pose_array[5][2])

        msg.r_pinky_ref.position.x = float(pose_array[6][0])
        msg.r_pinky_ref.position.z = float(pose_array[6][1])
        msg.r_pinky_ref.position.y = float(pose_array[6][2])

        msg.l_pinky_ref.position.x = float(pose_array[7][0])
        msg.l_pinky_ref.position.z = float(pose_array[7][1])
        msg.l_pinky_ref.position.y = float(pose_array[7][2])

        msg.r_index_ref.position.x = float(pose_array[8][0])
        msg.r_index_ref.position.z = float(pose_array[8][1])
        msg.r_index_ref.position.y = float(pose_array[8][2])

        msg.l_index_ref.position.x = float(pose_array[9][0])
        msg.l_index_ref.position.z = float(pose_array[9][1])
        msg.l_index_ref.position.y = float(pose_array[9][2])

        msg.r_thumb_ref.position.x = float(pose_array[10][0])
        msg.r_thumb_ref.position.z = float(pose_array[10][1])
        msg.r_thumb_ref.position.y = float(pose_array[10][2])

        msg.l_thumb_ref.position.x = float(pose_array[11][0])
        msg.l_thumb_ref.position.z = float(pose_array[11][1])
        msg.l_thumb_ref.position.y = float(pose_array[11][2])

        msg.r_hip_ref.position.x = float(pose_array[12][0])
        msg.r_hip_ref.position.z = float(pose_array[12][1])
        msg.r_hip_ref.position.y = float(pose_array[12][2])
        msg.r_hip_ref.orientation.y = float(pose_array[12][0]) # !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!

        msg.l_hip_ref.position.x = float(pose_array[13][0])
        msg.l_hip_ref.position.z = float(pose_array[13][1])
        msg.l_hip_ref.position.y = float(pose_array[13][2])
        #msg.l_hip_ref.orientation.y = float(angle_array[1])

        msg.r_knee_ref.position.x = float(pose_array[14][0])
        msg.r_knee_ref.position.z = float(pose_array[14][1])
        msg.r_knee_ref.position.y = float(pose_array[14][2])
        #msg.r_knee_ref.orientation.y = float(angle_array[2])

        msg.l_knee_ref.position.x = float(pose_array[15][0])
        msg.l_knee_ref.position.z = float(pose_array[15][1])
        msg.l_knee_ref.position.y = float(pose_array[15][2])
        #msg.l_knee_ref.orientation.y = float(angle_array[3])

        msg.r_ankle_ref.position.x = float(pose_array[16][0])
        msg.r_ankle_ref.position.z = float(pose_array[16][1])
        msg.r_ankle_ref.position.y = float(pose_array[16][2])
        #msg.r_ankle_ref.orientation.y = float(angle_array[4])

        msg.l_ankle_ref.position.x = float(pose_array[17][0])
        msg.l_ankle_ref.position.z = float(pose_array[17][1])
        msg.l_ankle_ref.position.y = float(pose_array[17][2])
        #msg.l_ankle_ref.orientation.y = float(pose_array[5])

        msg.r_heel_ref.position.x = float(pose_array[18][0])
        msg.r_heel_ref.position.z = float(pose_array[18][1])
        msg.r_heel_ref.position.y = float(pose_array[18][2])

        msg.l_heel_ref.position.x = float(pose_array[19][0])
        msg.l_heel_ref.position.z = float(pose_array[19][1])
        msg.l_heel_ref.position.y = float(pose_array[19][2])

        msg.r_foot_ref.position.x = float(pose_array[20][0])
        msg.r_foot_ref.position.z = float(pose_array[20][1])
        msg.r_foot_ref.position.y = float(pose_array[20][2])

        msg.l_foot_ref.position.x = float(pose_array[21][0])
        msg.l_foot_ref.position.z = float(pose_array[21][1])
        msg.l_foot_ref.position.y = float(pose_array[21][2])
        
        self.pub_skeleton.publish(msg)
        self.get_logger().info('INFORMATION PUBLISHED:\n'
                        '\n\n POSE COORDINATES:\n'
                        'Coordinates for right shoulder: [%f, %f, %f]\n'
                        'Coordinates for left shoulder: [%f, %f, %f]\n'
                        'Coordinates for right elbow: [%f, %f, %f]\n'
                        'Coordinates for left elbow: [%f, %f, %f]\n'
                        'Coordinates for right wrist: [%f, %f, %f]\n'
                        'Coordinates for left wrist: [%f, %f, %f]\n'
                        'Coordinates for right pinky finger: [%f, %f, %f]\n'
                        'Coordinates for left pinky finger: [%f, %f, %f]\n'
                        'Coordinates for right index finger: [%f, %f, %f]\n'
                        'Coordinates for left index finger: [%f, %f, %f]\n'
                        'Coordinates for right thumb finger: [%f, %f, %f]\n'
                        'Coordinates for left thumb finger: [%f, %f, %f]\n'
                        'Coordinates for right hip: [%f, %f, %f]\n'
                        'Coordinates for left hip: [%f, %f, %f]\n'
                        'Coordinates for right knee: [%f, %f, %f]\n'
                        'Coordinates for left knee: [%f, %f, %f]\n'
                        'Coordinates for right ankle: [%f, %f, %f]\n'
                        'Coordinates for left ankle: [%f, %f, %f]\n'
                        'Coordinates for right heel: [%f, %f, %f]\n'
                        'Coordinates for left heel: [%f, %f, %f]\n'
                        'Coordinates for right foot: [%f, %f, %f]\n'
                        'Coordinates for left foot: [%f, %f, %f]\n'
                        '\n\n KINEMATIC PARAMETERS:\n'
                        'Right hip angle: %f\n'
                        

                        % (msg.r_shoulder_ref.position.x, msg.r_shoulder_ref.position.y, msg.r_shoulder_ref.position.z,
                           msg.l_shoulder_ref.position.x, msg.l_shoulder_ref.position.y, msg.l_shoulder_ref.position.z,
                           msg.r_elbow_ref.position.x, msg.r_elbow_ref.position.y, msg.r_elbow_ref.position.z,
                           msg.l_elbow_ref.position.x, msg.l_elbow_ref.position.y, msg.l_elbow_ref.position.z,
                           msg.r_wrist_ref.position.x, msg.r_wrist_ref.position.y, msg.r_wrist_ref.position.z,
                           msg.l_wrist_ref.position.x, msg.l_wrist_ref.position.y, msg.l_wrist_ref.position.z,
                           msg.r_pinky_ref.position.x, msg.r_pinky_ref.position.y, msg.r_pinky_ref.position.z,
                           msg.l_pinky_ref.position.x, msg.l_pinky_ref.position.y, msg.l_pinky_ref.position.z,
                           msg.r_index_ref.position.x, msg.r_index_ref.position.y, msg.r_index_ref.position.z,
                           msg.l_index_ref.position.x, msg.l_index_ref.position.y, msg.l_index_ref.position.z,
                           msg.r_thumb_ref.position.x, msg.r_thumb_ref.position.y, msg.r_thumb_ref.position.z,
                           msg.l_thumb_ref.position.x, msg.l_thumb_ref.position.y, msg.l_thumb_ref.position.z,
                           msg.r_hip_ref.position.x, msg.r_hip_ref.position.y, msg.r_hip_ref.position.z,
                           msg.l_hip_ref.position.x, msg.l_hip_ref.position.y, msg.l_hip_ref.position.z,
                           msg.r_knee_ref.position.x, msg.r_knee_ref.position.y, msg.r_knee_ref.position.z,
                           msg.l_knee_ref.position.x, msg.l_knee_ref.position.y, msg.l_knee_ref.position.z,
                           msg.r_ankle_ref.position.x, msg.r_ankle_ref.position.y, msg.r_ankle_ref.position.z,
                           msg.l_ankle_ref.position.x, msg.l_ankle_ref.position.y, msg.l_ankle_ref.position.z,
                           msg.r_heel_ref.position.x, msg.r_heel_ref.position.y, msg.r_heel_ref.position.z,
                           msg.l_heel_ref.position.x, msg.l_heel_ref.position.y, msg.l_heel_ref.position.z,
                           msg.r_foot_ref.position.x, msg.r_foot_ref.position.y, msg.r_foot_ref.position.z,
                           msg.l_foot_ref.position.x, msg.l_foot_ref.position.y, msg.l_foot_ref.position.z,
                           msg.r_hip_ref.orientation.y)
                    )



class Config:
    def __init__(self, filename):
        if not os.path.exists(filename):
            print('File does not exist:', filename)
            quit()

        with open(filename) as f:
            config_data = yaml.safe_load(f)

        self.cam1_id = config_data['camera0']
        self.cam2_id = config_data['camera1']
        self.frame_width = config_data['frame_width']
        self.frame_height = config_data['frame_height']
        self.resize = config_data['view_resize']
        self.mono_frames = config_data['mono_cal_frames']
        self.stereo_frames = config_data['stereo_cal_frames']
        self.cooldown = config_data['cooldown']
        self.square_size = config_data['square_size']
        self.board_rows = config_data['board_rows']
        self.board_cols = config_data['board_cols']



class Camera:
    def __init__(self, id, name):
        self.id = id
        self.name = name
        self.frame_size = []
        self.fps = 30
        # time
        self.time = []
        
        #  CALIBRATION
        # intrinsics
        self.K = []
        self.dist = []
        # extrinsics
        self.R = []
        self.t = []
        # reprojection matrix
        self.P = []

        #   MEDIAPIPE POSE
        # mediapipe coordinates for each camera
        self.mp_kpts = {
            "R_shoulder" : [],  # 1
            "L_shoulder" : [],  # 2
            "R_elbow"    : [],  # 3
            "L_elbow"    : [],  # 4
            "R_wrist"    : [],  # 5 
            "L_wrist"    : [],  # 6
            "R_pinky"    : [],  # 7
            "L_pinky"    : [],  # 8 
            "R_index"    : [],  # 9
            "L_index"    : [],  # 10
            "R_thumb"    : [],  # 11
            "L_thumb"    : [],  # 12
            "R_hip"      : [],  # 13
            "L_hip"      : [],  # 14
            "R_knee"     : [],  # 15
            "L_knee"     : [],  # 16
            "R_ankle"    : [],  # 17
            "L_ankle"    : [],  # 18
            "R_heel"     : [],  # 19
            "L_heel"     : [],  # 20
            "R_foot"     : [],  # 21
            "L_foot"     : [],  # 22
        }

    # =============== CALIBRATION METHODS =============
    def take_mono_frames(self):
        """ This method saves the number of predefined pictures on the settings YAML file
            :return: 'mono_frames' folder with images
        """

        # Create directory if it doesn't exist already
        if not os.path.exists('mono_frames'):
            os.mkdir('mono_frames')

        # Open camera
        cap = cv.VideoCapture(self.id, cv.CAP_DSHOW)

        # Quit if the camera cannot be opened
        if not cap.isOpened():
            print('Error: Camera index out of range'
                  '\n\nPlease, make sure that the camera index is correct in the\n'
                  '"calibration_settings.yaml" file and rerun the program.')
            quit()

        # Set resolutions
        width = settings.frame_width
        height = settings.frame_height
        cap.set(3, width)
        cap.set(4, height)

        # Set additional parameters
        resize = settings.resize
        max_frames = settings.mono_frames
        cooldown_time = settings.cooldown
        cooldown = cooldown_time

        cur_frame = 0
        start = False
        font = cv.FONT_HERSHEY_COMPLEX

        while cap.isOpened():
            ret, frame = cap.read()
            if not ret:
                print('Camera not returning video data. Exiting...')
                quit()

            # For visualization purposes, reduce the size of the displayed video
            frame_small = cv.resize(frame, None, fx=1. / resize, fy=1. / resize)

            if not start:
                cv.putText(frame_small, "Make sure the camera shows the calibration pattern well", (50, 50), font, 0.7,
                           (0, 0, 255), 1)
                cv.putText(frame_small, "Press SPACEBAR to start collecting frames", (50, 100), font, 0.7,
                           (0, 0, 255), 1)

            if start:
                cooldown -= 1
                cv.putText(frame_small, "Cooldown: ", (50, 50), cv.FONT_HERSHEY_COMPLEX, 1,
                           (0, 0, 255), 1)
                cv.putText(frame_small, str(cooldown), (50, 150), cv.FONT_HERSHEY_COMPLEX, 3,
                           (255, 0, 0), 2)
                cv.putText(frame_small, "Num frames: ", (50, 200), cv.FONT_HERSHEY_COMPLEX, 1,
                           (0, 0, 255), 1)
                cv.putText(frame_small, str(cur_frame), (50, 300), cv.FONT_HERSHEY_COMPLEX, 3,
                           (255, 0, 0), 2)

                # save frame when cooldown reaches 0
                if cooldown <= 0:
                    savename = os.path.join('mono_frames', self.name + '_' + str(cur_frame) + '.png')
                    cv.imwrite(savename, frame)

                    print('Image saved! You have taken:', str(cur_frame), 'frames')
                    cur_frame += 1
                    cooldown = cooldown_time

            cv.imshow('Getting mono-frames for ' + str(self.name), frame_small)

            # keyboard actions
            k = cv.waitKey(1)

            if k == 27:  # ESC
                # if ESC is pressed at any time, the program will exit
                break
            elif k == 32:  # SPACEBAR
                # press spacebar to start data collection
                start = True

            # break out of the loop when enough number of frames have been saved
            if cur_frame == max_frames+1:
                break

        cv.destroyAllWindows()

    def get_intrinsics(self, cam_path):
        """ This method estimates the intrinsic parameters of the camera from the images
            stored in the 'mono_frames' folder
            :param:  camera_path
            :return: camera matrix and distortion coefficients
        """

        images_names = glob.glob(cam_path)

        # Read frames
        images = [cv.imread(imname, 1) for imname in images_names]

        # Criteria
        criteria = (cv.TERM_CRITERIA_EPS + cv.TERM_CRITERIA_MAX_ITER, 30, 0.001)

        # Creating vector to store vectors of 3D points for each checkerboard image
        objpoints = []
        # Creating vector to store vectors of 2D points for each checkerboard image
        imgpoints = []

        rows = settings.board_rows
        columns = settings.board_cols
        board = (rows, columns)
        resize = settings.resize
        world_scaling = settings.square_size

        # Get image shape from taken pictures
        img_shape = (images[0].shape[1], images[0].shape[0])

        # Defining the world coordinates for 3D points
        objp = np.zeros((1, board[0] * board[1], 3), np.float32)
        objp[0, :, :2] = np.mgrid[0:board[0], 0:board[1]].T.reshape(-1, 2)
        objp = world_scaling * objp

        # Extracting path of individual image stored in a given directory
        images = glob.glob(cam_path)

        for frame in images:

            img = cv.imread(frame)
            gray = cv.cvtColor(img, cv.COLOR_BGR2GRAY)

            # Find the checkerboard corners
            # If desired number of corners are found in the image then ret = true
            ret, corners = cv.findChessboardCorners(gray, board,
                                                    cv.CALIB_CB_ADAPTIVE_THRESH + cv.CALIB_CB_FAST_CHECK + cv.CALIB_CB_NORMALIZE_IMAGE)

            # CALIB_CB_ADAPTIVE_THRESH Use adaptive thresholding to convert the image to black and white
            # CALIB_CB_NORMALIZE_IMAGE Normalize the image gamma with #equalizeHist before applying fixed or adaptive thresholding
            # CALIB_CB_FAST_CHECK Run a fast check on the image that looks for chessboard corners, and shortcut the call if none is found

            if ret == True:

                conv_size = (11, 11)  # convolution size to improve corner detection

                # refining pixel coordinates for given 2d points
                corners2 = cv.cornerSubPix(gray, corners, conv_size, (-1, -1), criteria)

                # draw and display the corners
                cv.drawChessboardCorners(img, board, corners2, ret)
                # cv.putText(img, 'If detected points are poor, press "s" to skip this sample',
                #           (25, 25), cv.FONT_HERSHEY_COMPLEX, 1, (0, 0, 255), 1)

                frame_small = cv.resize(img, None, fx=1. / resize, fy=1. / resize)
                cv.imshow("Getting intrinsic parameters", frame_small)
                k = cv.waitKey(0)

                if k & 0xFF == ord('s'):
                    print('Skipping this sample...')
                    continue
                if k == 27:
                    print('Exiting...')
                    break

                objpoints.append(objp)
                imgpoints.append(corners2)

        cv.destroyAllWindows()

        # Calibration
        ret, mtx, dist, rvecs, tvecs = cv.calibrateCamera(objpoints, imgpoints, gray.shape[::-1], None, None)
        new_mtx, roi = cv.getOptimalNewCameraMatrix(mtx, dist, img_shape, 1, img_shape)

        dist = np.round(dist, decimals=3)
        new_mtx = np.round(new_mtx, decimals=3)

        print("\nReprojection error: {}".format(ret))

        self.K = new_mtx
        self.dist = dist

    def save_intrinsics(self):
        """ This method saves the parameters estimated with get_intrinsics
            :return .dat file with intrinsics:
        """
        # Create folder if it does not exist
        if not os.path.exists('camera_parameters'):
            os.mkdir('camera_parameters')

        out_filename = os.path.join('camera_parameters', self.name + '_intrinsics.dat')
        outf = open(out_filename, 'w')

        outf.write('K:\n')
        for i in self.K:
            for j in i:
                outf.write(str(j) + ' ')
            outf.write('\n')

        outf.write('dist:\n')
        for j in self.dist[0]:
            outf.write(str(j) + ' ')
        outf.write('\n')

    def read_intrinsics(self):
        """ This method reads the intrinsic parameters of the camera saved on a .dat file

        :return:
        """
        inf = open('./camera_parameters/' + str(self.name) + '_intrinsics.dat', 'r')
        K = []
        dist = []

        line = inf.readline()
        for _ in range(3):
            line = inf.readline().split()
            line = [float(en) for en in line]
            K.append(line)
        line = inf.readline()
        line = inf.readline().split()
        line = line = [float(en) for en in line]
        dist.append(line)
        inf.close()

        self.K = np.array(K)
        self.dist = np.array(dist)

    def save_extrinsics(self):
        """ This method saves the extrinsic values of individual cameras

        :param camera class Camera:
        :return:
        """
        # create folder if it does not exist
        if not os.path.exists('camera_parameters'):
            os.mkdir('camera_parameters')

        out_filename = os.path.join('camera_parameters', self.name + '_extrinsics.dat')
        outf = open(out_filename, 'w')

        outf.write('R:\n')
        for i in self.R:
            for j in i:
                outf.write(str(j) + ' ')
            outf.write('\n')

        outf.write('t:\n')
        for i in self.t:
            for j in i:
                outf.write(str(j) + ' ')
            outf.write('\n')

    def read_extrinsics(self):
        """ This method reads the extrinsic parameters of the camera saved on a .dat file

        :return:
        """
        inf = open('./camera_parameters/' + str(self.name) + '_extrinsics.dat', 'r')
        R = []
        t = []

        line = inf.readline()
        for _ in range(3):
            line = inf.readline().split()
            line = [float(en) for en in line]
            R.append(line)

        line = inf.readline()
        for _ in range(3):
            line = inf.readline().split()
            line = [float(en) for en in line]
            t.append(line)
        inf.close()

        self.R = np.array(R)
        self.t = np.array(t)

    def calculate_P_matrix(self):
        Rt = np.zeros((4, 4))
        Rt[:3, :3] = self.R
        Rt[:3, 3] = self.t.reshape(3)
        Rt[3, 3] = 1

        P = self.K @ Rt[:3, :]

        self.P = P

    def load_calibration(self):
        self.read_intrinsics()
        self.read_extrinsics()
        self.calculate_P_matrix()


class Skeleton:
    def __init__(self):

        self.publisher = PosePublisher()
        self.settings = Config("C:/dev/ros2_ws/src/pose_estimation/pose_estimation/settings.yaml")

        # Offsets for simulation

        # MediaPipe Pose solutions dictionary
        self.pose_keypoints = {
            "R_shoulder" : mp_pose.PoseLandmark.RIGHT_SHOULDER,     # 1
            "L_shoulder" : mp_pose.PoseLandmark.LEFT_SHOULDER,      # 2
            "R_elbow"    : mp_pose.PoseLandmark.RIGHT_ELBOW,        # 3
            "L_elbow"    : mp_pose.PoseLandmark.LEFT_ELBOW,         # 4
            "R_wrist"    : mp_pose.PoseLandmark.RIGHT_WRIST,        # 5 
            "L_wrist"    : mp_pose.PoseLandmark.LEFT_WRIST,         # 6
            "R_pinky"    : mp_pose.PoseLandmark.RIGHT_PINKY,        # 7
            "L_pinky"    : mp_pose.PoseLandmark.LEFT_PINKY,         # 8 
            "R_index"    : mp_pose.PoseLandmark.RIGHT_INDEX,        # 9
            "L_index"    : mp_pose.PoseLandmark.LEFT_INDEX,         # 10
            "R_thumb"    : mp_pose.PoseLandmark.RIGHT_THUMB,        # 11
            "L_thumb"    : mp_pose.PoseLandmark.LEFT_THUMB,         # 12
            "R_hip"      : mp_pose.PoseLandmark.RIGHT_HIP,          # 13
            "L_hip"      : mp_pose.PoseLandmark.LEFT_HIP,           # 14
            "R_knee"     : mp_pose.PoseLandmark.RIGHT_KNEE,         # 15
            "L_knee"     : mp_pose.PoseLandmark.LEFT_KNEE,          # 16
            "R_ankle"    : mp_pose.PoseLandmark.RIGHT_ANKLE,        # 17
            "L_ankle"    : mp_pose.PoseLandmark.LEFT_ANKLE,         # 18
            "R_heel"     : mp_pose.PoseLandmark.RIGHT_HEEL,         # 19
            "L_heel"     : mp_pose.PoseLandmark.LEFT_HEEL,          # 20
            "R_foot"     : mp_pose.PoseLandmark.RIGHT_FOOT_INDEX,   # 21
            "L_foot"     : mp_pose.PoseLandmark.LEFT_FOOT_INDEX,    # 22
        }

        # Dictionary containing final points
        #   It will be the same as camera.mp_kpts for 1 camera analysis
        #   It will be the dictionary containing the 3D points for 2 camera analysis
        self.pose_dict = {
            "R_shoulder" : [],  # 1
            "L_shoulder" : [],  # 2
            "R_elbow"    : [],  # 3
            "L_elbow"    : [],  # 4
            "R_wrist"    : [],  # 5 
            "L_wrist"    : [],  # 6
            "R_pinky"    : [],  # 7
            "L_pinky"    : [],  # 8 
            "R_index"    : [],  # 9
            "L_index"    : [],  # 10
            "R_thumb"    : [],  # 11
            "L_thumb"    : [],  # 12
            "R_hip"      : [],  # 13
            "L_hip"      : [],  # 14
            "R_knee"     : [],  # 15
            "L_knee"     : [],  # 16
            "R_ankle"    : [],  # 17
            "L_ankle"    : [],  # 18
            "R_heel"     : [],  # 19
            "L_heel"     : [],  # 20
            "R_foot"     : [],  # 21
            "L_foot"     : [],  # 22
        }

    # =============== CALIBRATION METHODS =============
    def take_stereo_frames(self, camera1, camera2):
        """ This method takes stereo-frames for two cameras
        :param camera1 class Camera:
        :param camera2 class Camera:
        :return 'stereo_frames' folder with stereo frames:
        """

        # Create directory if it doesn't exist already
        if not os.path.exists('stereo_frames'):
            os.mkdir('stereo_frames')

        # Get calibration settings
        resize = self.settings.resize
        cooldown_time = self.settings.cooldown
        cooldown = cooldown_time
        max_frames = self.settings.stereo_frames

        # Set additional parameters
        saved_count = 0
        start = False
        font = cv.FONT_HERSHEY_COMPLEX

        # Open cameras
        cap0 = cv.VideoCapture(camera1.id, cv.CAP_DSHOW)
        if not cap0.isOpened():
            print('Error: First camera index (camera0) out of range'
                  '\n\nPlease, make sure that the camera index is correct in the\n'
                  '"calibration_settings.yaml" file and rerun the program.')
            quit()

        cap1 = cv.VideoCapture(camera2.id, cv.CAP_DSHOW)
        if not cap1.isOpened():
            print('Error: Second camera index (camera1) out of range'
                  '\n\nPlease, make sure that the camera index is correct in the\n'
                  '"calibration_settings.yaml" file and rerun the program.')
            quit()

        # Set resolutions
        width = self.settings.frame_width
        height = self.settings.frame_height
        cap0.set(3, width)
        cap0.set(4, height)
        cap1.set(3, width)
        cap1.set(4, height)

        while cap0.isOpened() and cap1.isOpened():

            ret0, frame0 = cap0.read()
            ret1, frame1 = cap1.read()

            # Mirror cameras
            # frame0 = cv.flip(frame0, 1)
            # frame1 = cv.flip(frame1, 1)

            if not ret0 or not ret1:
                print('Cameras not returning video data. Exiting...')
                quit()

            frame0_small = cv.resize(frame0, None, fx=1. / resize, fy=1. / resize)
            frame1_small = cv.resize(frame1, None, fx=1. / resize, fy=1. / resize)

            if not start:
                cv.putText(frame0_small, "Make sure both cameras can see the calibration pattern well", (50, 50),
                           cv.FONT_HERSHEY_COMPLEX, 0.7, (0, 0, 255), 1)
                cv.putText(frame0_small, "Press SPACEBAR to start collecting frames", (50, 100),
                           cv.FONT_HERSHEY_COMPLEX, 0.7, (0, 0, 255), 1)

            if start:
                cooldown -= 1
                cv.putText(frame0_small, "Cooldown: ", (50, 50), cv.FONT_HERSHEY_COMPLEX, 1,
                           (0, 255, 0), 1)
                cv.putText(frame0_small, str(cooldown), (50, 150), cv.FONT_HERSHEY_COMPLEX, 3,
                           (255, 0, 0), 2)
                cv.putText(frame0_small, "Num frames: ", (50, 200), cv.FONT_HERSHEY_COMPLEX, 1,
                           (0, 255, 0), 1)
                cv.putText(frame0_small, str(saved_count), (50, 300), cv.FONT_HERSHEY_COMPLEX, 3,
                           (255, 0, 0), 2)

                # Save the frame when cooldown reaches 0
                if cooldown <= 0:
                    savename = os.path.join('stereo_frames', camera1.name + '_' + str(saved_count) + '.png')
                    cv.imwrite(savename, frame0)

                    savename = os.path.join('stereo_frames', camera2.name + '_' + str(saved_count) + '.png')
                    cv.imwrite(savename, frame1)

                    saved_count += 1
                    cooldown = cooldown_time

            # Show both images in the same window
            comb = np.concatenate((frame0_small, frame1_small), axis=1)
            comb = cv.resize(comb, (1536, 432))
            cv.imshow('Getting stereo frames', comb)

            # Keyboard actions
            k = cv.waitKey(1)
            if k == 27:  # ESC
                # if ESC is pressed at any time, the program will exit
                break
            elif k == 32:  # spacebar
                # press spacebar to start data collection
                start = True

            # Break out of the loop when the max number of frames has been reached
            if saved_count == max_frames+1:
                break

        cv.destroyAllWindows()

    def get_extrinsics(self, camera1, camera2):
        """ This method estimates the extrinsic parameters of the system taking camera1 as the origin

        :param camera1 class Camera:
        :param camera2 class Camera:
        :return:
        """

        prefix_c0 = os.path.join('stereo_frames', str(camera1.name) + '*')
        prefix_c1 = os.path.join('stereo_frames', str(camera2.name) + '*')

        # Read the stereo frames
        c0_images_names = sorted(glob.glob(prefix_c0))
        c1_images_names = sorted(glob.glob(prefix_c1))

        # Open images
        c0_images = [cv.imread(imname, 1) for imname in c0_images_names]
        c1_images = [cv.imread(imname, 1) for imname in c1_images_names]

        # change this if stereo calibration not good
        criteria = (cv.TERM_CRITERIA_EPS + cv.TERM_CRITERIA_MAX_ITER, 100, 0.001)

        # Calibration pattern settings
        rows = settings.board_rows
        columns = settings.board_cols
        world_scaling = settings.square_size

        # Coordinates of squares in the checkerboard world space
        objp = np.zeros((rows * columns, 3), np.float32)
        objp[:, :2] = np.mgrid[0:rows, 0:columns].T.reshape(-1, 2)
        objp = world_scaling * objp

        # Frame dimensions. Frames should be the same size.
        width = c0_images[0].shape[1]
        height = c0_images[0].shape[0]

        # Pixel coordinates of checkerboards
        imgpoints_left = []  # 2d points in image plane.
        imgpoints_right = []

        # Coordinates of the checkerboard in checkerboard world space.
        objpoints = []  # 3d point in real world space

        for frame0, frame1 in zip(c0_images, c1_images):

            gray1 = cv.cvtColor(frame0, cv.COLOR_BGR2GRAY)
            gray2 = cv.cvtColor(frame1, cv.COLOR_BGR2GRAY)
            c_ret1, corners1 = cv.findChessboardCorners(gray1, (rows, columns), None)
            c_ret2, corners2 = cv.findChessboardCorners(gray2, (rows, columns), None)

            if c_ret1 == True and c_ret2 == True:

                corners1 = cv.cornerSubPix(gray1, corners1, (11, 11), (-1, -1), criteria)
                corners2 = cv.cornerSubPix(gray2, corners2, (11, 11), (-1, -1), criteria)

                p0_c1 = corners1[0, 0].astype(np.int32)
                p0_c2 = corners2[0, 0].astype(np.int32)

                cv.putText(frame0, 'O', (p0_c1[0], p0_c1[1]), cv.FONT_HERSHEY_COMPLEX, 0.5, (0, 0, 255), 1)
                cv.drawChessboardCorners(frame0, (rows, columns), corners1, c_ret1)

                cv.putText(frame1, 'O', (p0_c2[0], p0_c2[1]), cv.FONT_HERSHEY_COMPLEX, 0.5, (0, 0, 255), 1)
                cv.drawChessboardCorners(frame1, (rows, columns), corners2, c_ret2)

                comb = np.concatenate((frame0, frame1), axis=1)
                comb = cv.resize(comb, (1536, 432))
                cv.imshow('Stereo Calibration', comb)

                k = cv.waitKey(0)

                if k & 0xFF == ord('s'):
                    print('skipping')
                    continue

                objpoints.append(objp)
                imgpoints_left.append(corners1)
                imgpoints_right.append(corners2)

        stereocalibration_flags = cv.CALIB_FIX_INTRINSIC
        ret, CM1, dist0, CM2, dist1, R, T, E, F = cv.stereoCalibrate(objpoints, imgpoints_left, imgpoints_right,
                                                                     camera1.K,
                                                                     camera1.dist,
                                                                     camera2.K, camera2.dist, (width, height),
                                                                     criteria=criteria,
                                                                     flags=stereocalibration_flags)

        print('RMSE: ', ret)

        camera1.R = np.eye(3, dtype=np.float32)
        camera1.t = np.array([0., 0., 0.]).reshape((3, 1))

        camera2.R = R
        camera2.t = T

        cv.destroyAllWindows()

    def calibration(self, camera1, camera2):
        # ___________________ INTRINSIC PARAMETERS _______________________
        # Get calibration frames for individual cameras and estimate and
        # save their intrinsic parameters
        #
        print('\nThe first step of the calibration process is to get the intrinsic parameters '
              'of both cameras. To do so, the specified number of mono frames on the settings file '
              'will be taken and saved on the folder "mono_frames". Then, the intrinsic parameters '
              'of the cameras will be estimated. Do you want to skip this step? (y/n)\n')
        ok = input('Answer: ')
        while ok not in ["n", "N", "y", "Y"]:
            print('Please, introduce a valid command:')
            ok = input('Answer: ')
        if ok == "n" or ok == "N":

            # CAMERA 0 ______________________________________________________
            print('\n\033[4mESTIMATING INTRINSIC PARAMETERS FOR THE FIRST CAMERA\033[0m')
            print('\nDo you want to take new pictures for camera0? (y/n)\n')
            pics = input('Answer: ')
            if pics == "N" or pics == "n":
                path = os.path.join('mono_frames', camera1.name + '_*')
                files = glob.glob(path)
                if not files:
                    print('No images found in:', path)
                    pics = "Y"
                else:
                    camera1.get_intrinsics(path)
            while pics not in ["n", "N", "y", "Y"]:
                print('Please, introduce a valid command:')
                pics = input('Answer: ')
            while pics == "y" or pics == "Y":
                print('\nOpening the first camera...')
                print('-> Make sure that the checkerboard pattern can be seen by the camera\n'
                      '-> To ensure a good calibration, take pictures from different angles\n'
                      'Press SPACEBAR to start saving frames.')
                camera1.take_mono_frames()
                print('\nA new window should have opened showing the checkerboard pattern with the detected corners.\n'
                      'If the detections are poor, skip the sample by pressing "s".\n'
                      'Otherwise, press any other key.\n'
                      'Anytime you want to close any emergent window, press ESC\n')
                path = os.path.join('mono_frames', camera1.name + '_*')
                camera1.get_intrinsics(path)
                print('Do you want to repeat the process? (y/n)\n'
                      '(You should aim for a RMSE < 0.5)\n')
                rep = input('Answer: ')
                if rep == "N" or rep == "n":
                    break
                elif rep == "Y" or rep == "y":
                    pics == 'Y'
                elif rep not in ["n", "N", "y", "Y"]:
                    print('Please, introduce a valid command:')
                    rep = input('Answer: ')
            print('___' * 25, '\n')
            print("K matrix =\n", camera1.K)
            print("\nDistortion coefficients =\n", camera1.dist)

            # Save parameters
            camera1.save_intrinsics()
            print('\nIntrinsic parameters correctly saved for the first camera in "camera_parameters" folder.')
            print('===' * 25, '\n')

            # CAMERA 1 ________________________________________________________

            print('\033[4mESTIMATING INTRINSIC PARAMETERS FOR THE SECOND CAMERA\033[0m')
            print('\nDo you want to take new pictures for camera1? (y/n)\n')
            pics = input('Answer: ')
            if pics == "N" or pics == "n":
                path = os.path.join('mono_frames', camera2.name + '_*')
                files = glob.glob(path)
                if not files:
                    print('No images found in:', path)
                    pics = "Y"
                else:
                    camera2.get_intrinsics(path)
            while pics not in ["n", "N", "y", "Y"]:
                print('Please, introduce a valid command:')
                pics = input('Answer: ')
            while pics == "y" or pics == "Y":
                print('\nOpening the second camera...')
                print('-> Make sure that the checkerboard pattern can be seen by the camera.\n'
                      '-> To ensure a good calibration, take pictures from different angles\n'
                      'Press SPACEBAR to start saving frames.')
                camera2.take_mono_frames()
                print('\nA new window should have opened showing the checkerboard pattern with the detected corners.\n'
                      'If the detections are poor, skip the sample by pressing "s".\n'
                      'Otherwise, press any other key.\n'
                      'Anytime you want to close any emergent window, press ESC\n')
                path = os.path.join('mono_frames', camera2.name + '_*')
                camera2.get_intrinsics(path)
                print('Do you want to repeat the process? (y/n)\n'
                      '(You should aim for a RMSE < 0.5)\n')
                rep = input('Answer: ')
                if rep == "N" or rep == "n":
                    break
                elif rep == "Y" or rep == "y":
                    pics == 'Y'
                elif rep not in ["n", "N", "y", "Y"]:
                    print('Please, introduce a valid command:')
                    rep = input('Answer: ')
            print('___' * 25, '\n')
            print("K matrix =\n", camera2.K)
            print("\nDistortion coefficients \n", camera2.dist)

            # Save parameters
            camera2.save_intrinsics()
            print('\nIntrinsic parameters correctly saved for the second camera in "camera_parameters" folder.\n')
            print('Calibration for individual cameras completed!')
            print('===' * 25, '\n')

        # ___________________ STEREO-CALIBRATION _______________________
        print('\n\033[4mSTEREO-CALIBRATION\033[0m')
        print('\nThe next step of the calibration process is to perform the stereo-calibration.\n\n'
              'Some pictures will be taken simultaneously with both cameras to estimate the '
              'extrinsic parameters that define the geometry of the system. '
              'Finally, the parameters will be saved on the "camera_parameters" folder.\n'
              '\nDo you want to take the pictures? (y/n)\n')
        ok = input('Answer: ')
        if ok == "N" or ok == "n":
            path1 = os.path.join('stereo_frames', camera1.name + '_*')
            files1 = glob.glob(path1)
            path2 = os.path.join('stereo_frames', camera2.name + '_*')
            files2 = glob.glob(path2)
            if not files1 or not files2:
                print('No stereo-images found')
                ok = "Y"
            else:
                # Import intrinsic parameters
                camera1.read_intrinsics()
                camera2.read_intrinsics()

                # Get extrinsics
                self.get_extrinsics(camera1, camera2)

        while ok not in ["n", "N", "y", "Y"]:
            print('Please, introduce a valid command:')
            ok = input('Answer: ')
        # Taking and saving stereo-frames
        while ok == "y" or ok == "Y":
            print('\nOpening both cameras...')
            print('This may take some time.\n'
                  '\nPress SPACEBAR to start the collection of frames.')
            self.take_stereo_frames(camera1, camera2)

            # Import intrinsic parameters
            camera1.read_intrinsics()
            camera2.read_intrinsics()

            # Get extrinsics
            self.get_extrinsics(camera1, camera2)

            print('Do you want to repeat the process? (y/n)\n'
                  '(You should aim for a RMSE < 0.5)\n')
            rep = input('Answer: ')
            if rep == "N" or rep == "n":
                ok == 'n'
                break
            elif rep == "Y" or rep == "y":
                ok == 'y'
            elif rep not in ["n", "N", "y", "Y"]:
                print('Please, introduce a valid command:')
                rep = input('Answer: ')

        # Save extrinsics
        camera1.save_extrinsics()
        camera2.save_extrinsics()
        print('\nExtrinsic parameters correctly saved!')

    def stereo_calibration(self, camera1, camera2):
        print('\n\033[4mSTEREO-CALIBRATION\033[0m')
        print('\nThe next step of the calibration process is to perform the stereo-calibration.\n\n'
              'Some pictures will be taken simultaneously with both cameras to estimate the '
              'extrinsic parameters\nthat define the geometry of the system. '
              'Finally, the parameters will be saved\non the "camera_parameters" folder.\n'
              '\nDo you want to take the pictures? (y/n)\n')
        ok = input('Answer: ')
        if ok == "N" or ok == "n":
            path1 = os.path.join('stereo_frames', camera1.name + '_*')
            files1 = glob.glob(path1)
            path2 = os.path.join('stereo_frames', camera2.name + '_*')
            files2 = glob.glob(path2)
            if not files1 or not files2:
                print('No stereo-images found')
                ok = "Y"
            else:
                # Import intrinsic parameters
                camera1.read_intrinsics()
                camera2.read_intrinsics()

                # Get extrinsics
                self.get_extrinsics(camera1, camera2)

        while ok not in ["n", "N", "y", "Y"]:
            print('Please, introduce a valid command:')
            ok = input('Answer: ')
        # Taking and saving stereo-frames
        while ok == "y" or ok == "Y":
            print('\nOpening both cameras...')
            print('This may take some time.\n'
                  '\nPress SPACEBAR to start the collection of frames.')
            self.take_stereo_frames(camera1, camera2)

            # Import intrinsic parameters
            camera1.read_intrinsics()
            camera2.read_intrinsics()

            # Get extrinsics
            self.get_extrinsics(camera1, camera2)

            print('Do you want to repeat the process? (y/n)\n'
                  '(You should aim for a RMSE < 0.5)\n')
            rep = input('Answer: ')
            if rep == "N" or rep == "n":
                break
            elif rep == "Y" or rep == "y":
                ok == 'y'
            elif rep not in ["n", "N", "y", "Y"]:
                print('Please, introduce a valid command:')
                rep = input('Answer: ')

        # Save extrinsics
        camera1.save_extrinsics()
        camera2.save_extrinsics()
        print('\nExtrinsic parameters correctly saved!')

    # ============= MEDIAPIPE POSE ANALYSIS ===========
    def analysisVideo(self, video_path):
        """ MediaPipe Pose analysis of a recorded video.

        :param video_path:
        :return:
        """
        # Load video
        video = cv.VideoCapture(video_path)

        # Get height and width
        w = int(video.get(cv.CAP_PROP_FRAME_WIDTH))
        h = int(video.get(cv.CAP_PROP_FRAME_HEIGHT))

        # Get rescaled height and width for visualization purposes
        scale_percent = 30  # percent of original size
        width = int(w * scale_percent / 100)
        height = int(h * scale_percent / 100)

        # Define extra parameters
        frame_num = 0  # set counter to 0

        # Create body keypoint detector object
        pose = mp_pose.Pose(model_complexity=2, min_detection_confidence=0.7, min_tracking_confidence=0.5)

        # Create folder if it does not exist for pose_results
        if not os.path.exists('pose_results'):
            os.mkdir('pose_results')

        # Prepare video to save MediaPipe Pose results
        trial = input('Write savename for 2D pose results (.mat and .mp4): ')

        # Create VideoWriter objects
        output_name = os.path.join('pose_results', f'{trial}_pose_video.mp4')
        result = cv.VideoWriter(output_name,
                                cv.VideoWriter_fourcc(*'MP4V'),
                                30, (w, h))
        

        # Run MediaPipe Pose
        while True:

            # Keep track of current frame
            frame_num += 1

            # Read video
            ret, frame = video.read()
            if not ret:
                print('Error reading video')
                break

            # Turn BGR image to RGB
            frame = cv.cvtColor(frame, cv.COLOR_BGR2RGB)
            # Mark image as not writeable to improve performance
            frame.flags.writeable = False
            # Process results
            results = pose.process(frame)
            # Reverse changes
            frame.flags.writeable = True
            frame = cv.cvtColor(frame, cv.COLOR_RGB2BGR)

            # Draw results
            mp_drawing.draw_landmarks(frame, results.pose_landmarks, mp_pose.POSE_CONNECTIONS,
                                      landmark_drawing_spec=mp_drawing_style.get_default_pose_landmarks_style())

            cv.namedWindow('MediaPipe results', cv.WINDOW_NORMAL & cv.WINDOW_KEEPRATIO)
            cv.resizeWindow('MediaPipe results', width, height)
            cv.imshow('MediaPipe results', frame)

            # Save video with results
            result.write(frame)

            # Check for keypoints detection
            pose_array = []
            if results.pose_landmarks:
                print('_' * 25)
                print('Results for frame: ', frame_num)
                pos = 0
                for kpt in self.pose_keypoints:

                    x = float(results.pose_landmarks.landmark[self.pose_keypoints[kpt]].x)
                    y = 1 - float(results.pose_landmarks.landmark[self.pose_keypoints[kpt]].y)
                    z = float(results.pose_landmarks.landmark[self.pose_keypoints[kpt]].z)
                    
                    xyz = np.array([x, y, z])

                    pose_array.append(np.array([x, y, z], dtype=float))
                    self.pose_dict[kpt].append([x, y, z])
                    
                    pos += 1

                    #print('\n', str(kpt), ' - ', xyz)
            else:  
                pose_array = np.array([[-1, -1, -1]] * len(self.pose_keypoints))
                for kpt in self.pose_keypoints:
                    self.pose_dict[kpt].append([-1, -1, -1])

            # ROS2: Publish message
            self.publisher.publish_skeleton(pose_array)

            k = cv.waitKey(1)
            if k == 27:  # Press ESC to quit
                break
      
        cv.destroyAllWindows()
        self.publisher.destroy_node()
        result.release()

        # Save dictionaries as .mat files on the results directory
        savename = os.path.join('pose_results', f'{trial}_pose_2d.mat')
        scipy.io.savemat(savename, self.pose_dict)



        print('Process finished!')

    def analysisStereoVideo(self, path0, path1, camera1, camera2):
        # Load video
        video0 = cv.VideoCapture(path0)
        video1 = cv.VideoCapture(path1)

        # Get height and width
        w0 = int(video0.get(cv.CAP_PROP_FRAME_WIDTH))
        h0 = int(video0.get(cv.CAP_PROP_FRAME_HEIGHT))
        w1 = int(video1.get(cv.CAP_PROP_FRAME_WIDTH))
        h1 = int(video1.get(cv.CAP_PROP_FRAME_HEIGHT))

        # Keep track of frame number
        frame_num = 0

        # Initiate matplot figure
        fig = plt.figure()
        ax = fig.add_subplot(111, projection='3d')
        plt.ion()

        # Create folder if it does not exist for pose_results
        if not os.path.exists('pose_results'):
            os.mkdir('pose_results')

        # Prepare video to save MediaPipe Pose results
        trial = input('Write savename for 3D pose results (.mat and .mp4): ')
        # Create VideoWriter objects
        output_name0 = os.path.join('pose_results', f'{trial}_pose_video_L.mp4')
        result0 = cv.VideoWriter(output_name0,
                                cv.VideoWriter_fourcc(*'MP4V'),
                                30, (self.settings.frame_width, self.settings.frame_height))
        
        output_name1 = os.path.join('pose_results', f'{trial}_pose_video_R.mp4')
        result1 = cv.VideoWriter(output_name1,
                                cv.VideoWriter_fourcc(*'MP4V'),
                                30, (self.settings.frame_width, self.settings.frame_height))


        # Create body keypoint detector objects
        pose0 = mp_pose.Pose(model_complexity=2, min_detection_confidence=0.7, min_tracking_confidence=0.5)
        pose1 = mp_pose.Pose(model_complexity=2, min_detection_confidence=0.7, min_tracking_confidence=0.5)

        while True:
            frame_num += 1

            # Read frames from stream
            ret0, frame0 = video0.read()
            ret1, frame1 = video1.read()

            if not ret0 or not ret1:
                print('Error reading video')
                break

            # Turn BGR image to RGB
            frame0 = cv.cvtColor(frame0, cv.COLOR_BGR2RGB)
            frame1 = cv.cvtColor(frame1, cv.COLOR_BGR2RGB)

            # Mark images as not writeable
            frame0.flags.writeable = False
            frame1.flags.writeable = False

            # Process results
            results0 = pose0.process(frame0)
            results1 = pose1.process(frame1)

            # Reverse changes
            frame0.flags.writeable = True
            frame1.flags.writeable = True
            frame0 = cv.cvtColor(frame0, cv.COLOR_RGB2BGR)
            frame1 = cv.cvtColor(frame1, cv.COLOR_RGB2BGR)

            # Display results
            mp_drawing.draw_landmarks(frame0, results0.pose_landmarks, mp_pose.POSE_CONNECTIONS,
                                    landmark_drawing_spec=mp_drawing_style.get_default_pose_landmarks_style())
            mp_drawing.draw_landmarks(frame1, results1.pose_landmarks, mp_pose.POSE_CONNECTIONS,
                                    landmark_drawing_spec=mp_drawing_style.get_default_pose_landmarks_style())

                        
            comb = np.concatenate((frame0, frame1), axis=0)
            comb = cv.resize(comb, (699, 952))
            cv.imshow('MediaPipe results', comb)

            # Save videos
            result0.write(frame0)
            result1.write(frame1)

            # Check for keypoint detection
            frame0_array = []
            frame1_array = []
            if results0.pose_landmarks and results1.pose_landmarks:
                print('Results for frame: ', frame_num)
                for kpt in self.pose_keypoints:

                    x0 = int(results0.pose_landmarks.landmark[self.pose_keypoints[kpt]].x * w0)
                    y0 = int(results0.pose_landmarks.landmark[self.pose_keypoints[kpt]].y * h0)
                    frame0_array.append(np.array([x0, y0], dtype=float))

                    camera1.mp_kpts[kpt].append([x0, y0])
                    
                    
                    x1 = int(results1.pose_landmarks.landmark[self.pose_keypoints[kpt]].x * w1)
                    y1 = int(results1.pose_landmarks.landmark[self.pose_keypoints[kpt]].y * h1)
                    frame1_array.append(np.array([x1, y1], dtype=float))

                    camera2.mp_kpts[kpt].append([x1, y1])
                    

            else:
                # Add [-1,-1] to dictionaries if no keypoints are found
                frame0_array = np.array([[-1, -1]] * len(self.pose_keypoints))
                frame1_array = np.array([[-1, -1]] * len(self.pose_keypoints))

            # Transform 2D points to 3D points
            pose_array = []
            for kpt, uv1, uv2 in zip(self.pose_keypoints, frame0_array, frame1_array):
                p3d = np.array([-1, -1, -1])
                if uv1[0] != -1 and uv2[0] != -1:
                    p4d = cv.triangulatePoints(camera1.P, camera2.P, uv1, uv2)  # calculate 3d position of keypoints
                    p3d = (p4d[:3, 0] / p4d[3, 0]).T

                self.pose_dict[kpt].append([p3d[0], p3d[1], p3d[2]])
                pose_array.append(np.array([p3d[0], p3d[2], 1080-p3d[1]], dtype=float))  # Y and Z are exchanged


            # ROS2: Publish message
            self.publisher.publish_skeleton(pose_array)

            # Plot pose
            ax.cla()

            points_3d = pose_array.copy()
            points_3d = np.array(points_3d)

            ax.plot(points_3d[12:22:2, 0], points_3d[12:22:2, 1], points_3d[12:22:2, 2], 'b--.')  # Pierna derecha        
            ax.plot([points_3d[20, 0], points_3d[16, 0]], [points_3d[20, 1], points_3d[16, 1]], [points_3d[20, 2], points_3d[16, 2]], 'b--.')

            ax.plot(points_3d[13:23:2, 0], points_3d[13:23:2, 1], points_3d[13:23:2, 2], 'b--.')  # Pierna izquierda
            ax.plot([points_3d[21, 0], points_3d[17, 0]], [points_3d[21, 1], points_3d[17, 1]], [points_3d[21, 2], points_3d[17, 2]], 'b--.')

            ax.plot(points_3d[12:14, 0], points_3d[12:14, 1], points_3d[12:14, 2], 'g--.')  # Cadera
            ax.plot(points_3d[:2, 0], points_3d[:2, 1], points_3d[:2, 2], 'g--.')  # Hombros

            ax.plot([points_3d[12, 0], points_3d[0, 0]], [points_3d[12, 1], points_3d[0, 1]], [points_3d[12, 2], points_3d[0, 2]], 'g--.')  # Lado derecho
            ax.plot([points_3d[1, 0], points_3d[13, 0]], [points_3d[1, 1], points_3d[13, 1]], [points_3d[1, 2], points_3d[13, 2]], 'g--.')  # Lado izquierdo

            #ax.plot(points_3d[0:9:2, 0], points_3d[0:9:2, 1], points_3d[0:9:2, 2], 'r--.')  # Brazo derecho
            #ax.plot(points_3d[1:10:2, 0], points_3d[1:10:2, 1], points_3d[1:10:2, 2], 'r--.')  # Brazo izquierdo

            plt.show()

            # Resto del código para personalizar la visualización y ajustar los límites
            ax.set_xlim([-2000, 500])
            ax.set_ylim([2000, 4000])
            ax.set_zlim([0, 1750])
            ax.set_title('3D MediaPipe Detections')
            ax.set_xlabel('Width (mm)')
            ax.set_ylabel('Depth (mm)')
            ax.set_zlabel('Height (mm)')
            ax.view_init(30, 0)

            #########################################

            print('\n')
            print('=' * 30)
            print('\n\n')

            k = cv.waitKey(1)
            if k & 0xFF == 27: break  # 27 is ESC key

        cv.destroyAllWindows()
        result0.release()
        result1.release()
        self.publisher.destroy_node()

        # Save dictionaries as .mat files on the results directory
        if not os.path.exists('pose_results'):
            os.mkdir('pose_results')

        savename = os.path.join('pose_results', f'{trial}_pose_3d.mat')
        scipy.io.savemat(savename, self.pose_dict)

    def analysis1Camera(self, camera):
        """ Code for MediaPipe Pose using one camera

        :param camera class Camera:
        :return video with MP results:
        """

        # Input video stream
        cap = cv.VideoCapture(camera.id, cv.CAP_DSHOW)

        # Set camera resolution
        cap.set(3, self.settings.frame_width)
        cap.set(4, self.settings.frame_height)

        # Create body keypoint detector object
        pose = mp_pose.Pose(model_complexity=2, min_detection_confidence=0.7, min_tracking_confidence=0.5)

        # Keep track of frame number
        frame_num = 0
        angle_array = np.zeros((7,), dtype=float)

        # Prepare video to save MediaPipe Pose results
        #mp_vid = self.save_video_results()

        while True:

            # Keep track of current frame
            frame_num += 1

            # Read frame from stream
            ret, frame = cap.read()

            if not ret:
                print('Error getting video stream')
                break

            # Turn BGR image to RGB
            frame = cv.cvtColor(frame, cv.COLOR_BGR2RGB)
            # Mark image as not writeable to improve performance
            frame.flags.writeable = False
            # Process results
            results = pose.process(frame)
            # Reverse changes
            frame.flags.writeable = True
            frame = cv.cvtColor(frame, cv.COLOR_RGB2BGR)

            # Display results
            mp_drawing.draw_landmarks(frame, results.pose_landmarks, mp_pose.POSE_CONNECTIONS,
                                      landmark_drawing_spec=mp_drawing_style.get_default_pose_landmarks_style())

            frame_small = cv.resize(frame, (960, 540))
            cv.imshow('MediaPipe results', frame_small)

            # Save video with results
            #mp_vid.write(frame)

            # Check for keypoints detection
            pose_array = []
            if results.pose_landmarks:
                print('_' * 25)
                print('Results for frame: ', frame_num)
                for kpt in self.pose_keypoints:

                    x = float(results.pose_landmarks.landmark[self.pose_keypoints[kpt]].x)
                    y = 1 - float(results.pose_landmarks.landmark[self.pose_keypoints[kpt]].y)
                    #z = float(results.pose_landmarks.landmark[self.pose_keypoints[kpt]].z)
                    z = 0
                    xyz = np.array([x, y, z])

                    pose_array.append(np.array([x, y, z], dtype=float))
                    self.pose_dict[kpt].append([x, y, z])

                    # ROS2: Publish point
                    #self.publisher.publish_point(x, y, z)

            else:  
                pose_array = np.array([[-1, -1, -1]] * len(self.pose_keypoints))
                for kpt in self.pose_keypoints:
                    self.pose_dict[kpt].append([-1, -1, -1])

            # ROS2: Publish message
            self.publisher.publish_skeleton(pose_array)

            k = cv.waitKey(1)
            if k & 0xFF == 27:  # ESC
                break

        cv.destroyAllWindows()
        self.publisher.destroy_node()
        cap.release()

    def analysis2Cameras(self, camera1, camera2):
        
        # Input video streams
        cap0 = cv.VideoCapture(camera1.id, cv.CAP_DSHOW)
        cap1 = cv.VideoCapture(camera2.id, cv.CAP_DSHOW)
        caps = [cap0, cap1]

        # Set camera resolutions
        for cap in caps:
            cap.set(3, self.settings.frame_width)
            cap.set(4, self.settings.frame_height)

        # Create body keypoint detector objects
        pose0 = mp_pose.Pose(model_complexity=2, min_detection_confidence=0.7, min_tracking_confidence=0.5)
        pose1 = mp_pose.Pose(model_complexity=2, min_detection_confidence=0.7, min_tracking_confidence=0.5)

        # Keep track of frame number
        frame_num = 0
        angle_array = [0, 0, 0, 0, 0, 0]

        # Initiate matplot figure
        fig = plt.figure()
        ax = fig.add_subplot(111, projection='3d')
        plt.ion()

        while True:
            frame_num += 1

            # Read frames from stream
            ret0, frame0 = cap0.read()
            ret1, frame1 = cap1.read()

            if not ret0 or not ret1:
                print('Error reading video')
                break

            # Turn BGR image to RGB
            frame0 = cv.cvtColor(frame0, cv.COLOR_BGR2RGB)
            frame1 = cv.cvtColor(frame1, cv.COLOR_BGR2RGB)

            # Mark images as not writeable
            frame0.flags.writeable = False
            frame1.flags.writeable = False

            # Process results
            results0 = pose0.process(frame0)
            results1 = pose1.process(frame1)

            # Reverse changes
            frame0.flags.writeable = True
            frame1.flags.writeable = True
            frame0 = cv.cvtColor(frame0, cv.COLOR_RGB2BGR)
            frame1 = cv.cvtColor(frame1, cv.COLOR_RGB2BGR)

            # Display results
            mp_drawing.draw_landmarks(frame0, results0.pose_landmarks, mp_pose.POSE_CONNECTIONS,
                                    landmark_drawing_spec=mp_drawing_style.get_default_pose_landmarks_style())
            mp_drawing.draw_landmarks(frame1, results1.pose_landmarks, mp_pose.POSE_CONNECTIONS,
                                    landmark_drawing_spec=mp_drawing_style.get_default_pose_landmarks_style())

            comb = np.concatenate((frame0, frame1), axis=0)
            comb = cv.resize(comb, (699, 952))
            cv.imshow('MediaPipe results', comb)

            # Save videos with results
            # mp_vid.write(comb)

            # Check for keypoint detection
            frame0_array = []
            frame1_array = []
            if results0.pose_landmarks and results1.pose_landmarks:
                print('Results for frame: ', frame_num)
                for kpt in self.pose_keypoints:

                    x0 = int(results0.pose_landmarks.landmark[self.pose_keypoints[kpt]].x * frame0.shape[1])
                    y0 = int(results0.pose_landmarks.landmark[self.pose_keypoints[kpt]].y * frame0.shape[0])
                    xy0 = np.array([x0, y0])
                    frame0_array.append(np.array([x0, y0], dtype=float))

                    camera1.mp_kpts[kpt].append([x0, y0])


                    x1 = int(results1.pose_landmarks.landmark[self.pose_keypoints[kpt]].x * frame1.shape[1])
                    y1 = int(results1.pose_landmarks.landmark[self.pose_keypoints[kpt]].y * frame1.shape[0])
                    xy1 = np.array([x1, y1])
                    frame1_array.append(np.array([x1, y1], dtype=float))

                    camera2.mp_kpts[kpt].append([x1, y1])


            else:
                # Add [-1,-1] to dictionaries if no keypoints are found
                frame0_array = np.array([[-1, -1]] * len(self.pose_keypoints))
                frame1_array = np.array([[-1, -1]] * len(self.pose_keypoints))

            # Transform 2D points to 3D points
            pose_array =[]
            for kpt, uv1, uv2 in zip(self.pose_keypoints, frame0_array, frame1_array):
                p3d = np.array([-1, -1, -1])
                if uv1[0] != -1 and uv2[0] != -1:
                    p4d = cv.triangulatePoints(camera1.P, camera2.P, uv1, uv2)  # calculate 3d position of keypoints
                    p3d = (p4d[:3, 0] / p4d[3, 0]).T

                self.pose_dict[kpt].append([p3d[0], p3d[1], p3d[2]])
                pose_array.append(np.array([p3d[0], p3d[2], 1080-p3d[1]], dtype=float))  # Y and Z are exchanged


            # ROS2: Publish message
            self.publisher.publish_skeleton(pose_array)
            
            # Plot pose
            ax.cla()

            points_3d = pose_array.copy()
            points_3d = np.array(points_3d)

            ax.plot(points_3d[12:22:2, 0], points_3d[12:22:2, 1], points_3d[12:22:2, 2], 'b--.')  # Pierna derecha        
            ax.plot([points_3d[20, 0], points_3d[16, 0]], [points_3d[20, 1], points_3d[16, 1]], [points_3d[20, 2], points_3d[16, 2]], 'b--.')

            ax.plot(points_3d[13:23:2, 0], points_3d[13:23:2, 1], points_3d[13:23:2, 2], 'b--.')  # Pierna izquierda
            ax.plot([points_3d[21, 0], points_3d[17, 0]], [points_3d[21, 1], points_3d[17, 1]], [points_3d[21, 2], points_3d[17, 2]], 'b--.')

            ax.plot(points_3d[12:14, 0], points_3d[12:14, 1], points_3d[12:14, 2], 'g--.')  # Cadera
            ax.plot(points_3d[:2, 0], points_3d[:2, 1], points_3d[:2, 2], 'g--.')  # Hombros

            ax.plot([points_3d[12, 0], points_3d[0, 0]], [points_3d[12, 1], points_3d[0, 1]], [points_3d[12, 2], points_3d[0, 2]], 'g--.')  # Lado derecho
            ax.plot([points_3d[1, 0], points_3d[13, 0]], [points_3d[1, 1], points_3d[13, 1]], [points_3d[1, 2], points_3d[13, 2]], 'g--.')  # Lado izquierdo

            ax.plot(points_3d[0:9:2, 0], points_3d[0:9:2, 1], points_3d[0:9:2, 2], 'r--.')  # Brazo derecho
            ax.plot(points_3d[1:10:2, 0], points_3d[1:10:2, 1], points_3d[1:10:2, 2], 'r--.')  # Brazo izquierdo

            plt.show()

            # Resto del código para personalizar la visualización y ajustar los límites
            ax.set_xlim([-2000, 500])
            ax.set_ylim([1500, 3000])
            ax.set_zlim([0, 1750])
            ax.set_title('3D MediaPipe Detections')
            ax.set_xlabel('Width (mm)')
            ax.set_ylabel('Depth (mm)')
            ax.set_zlabel('Height (mm)')
            ax.view_init(30, 0)

            #########################################

            print('\n')
            print('=' * 30)
            print('\n\n')

            k = cv.waitKey(1)
            if k & 0xFF == 27: break  # 27 is ESC key

        cv.destroyAllWindows()
        self.publisher.destroy_node()
        for cap in caps:
            cap.release()    

    # ==================== RECORD =====================
    def StereoRecord(self, camera1, camera2):
        # Input video streams
        cap0 = cv.VideoCapture(camera1.id, cv.CAP_DSHOW) # DirectShow API Windows
        cap1 = cv.VideoCapture(camera2.id, cv.CAP_DSHOW)
        caps = [cap0, cap1]


        # Set camera resolutions
        for cap in caps:
            cap.set(3, self.settings.frame_width)
            cap.set(4, self.settings.frame_height)

        # Create folder for video_results
        if not os.path.exists('video_results'):
            os.mkdir('video_results')

        trial = input('Write trial number: ')
        # Create VideoWriter objects
        output_name0 = os.path.join('video_results', f'{trial}_L.mp4')
        result0 = cv.VideoWriter(output_name0,
                                cv.VideoWriter_fourcc(*'MP4V'),
                                30, (self.settings.frame_width, self.settings.frame_height))
        
        output_name1 = os.path.join('video_results', f'{trial}_R.mp4')
        result1 = cv.VideoWriter(output_name1,
                                cv.VideoWriter_fourcc(*'MP4V'),
                                30, (self.settings.frame_width, self.settings.frame_height))

        

        while True:

            # Read frames from stream
            ret0, frame0 = cap0.read()
            ret1, frame1 = cap1.read()

            if not ret0 or not ret1:
                print('Error reading video')
                break
            
            # Save videos
            result0.write(frame0)
            result1.write(frame1)


            # Show video results
            comb = np.concatenate((frame0, frame1), axis=1)
            comb = cv.resize(comb, (1280, 360))
            cv.imshow('Video results', comb)

            k = cv.waitKey(1)
            if k & 0xFF == 27: break  # 27 is ESC key


        cv.destroyAllWindows()
        for cap in caps:
            cap.release()
        result0.release()
        result1.release()

        print("Videos saved successfully!\n")  
            
    def check_fps(self, camera):
        # Start default camera
        video = cv.VideoCapture(camera.id)
    
        # Find OpenCV version
        (major_ver, minor_ver, subminor_ver) = (cv.__version__).split('.')
    
        # With webcam get(CV_CAP_PROP_FPS) does not work.
        # Let's see for ourselves.
    
        if int(major_ver)  < 3 :
            fps = video.get(cv.cv.CV_CAP_PROP_FPS)
            print("Frames per second using video.get(cv.cv.CV_CAP_PROP_FPS): {0}".format(fps))
        else :
            fps = video.get(cv.CAP_PROP_FPS)
            print("Frames per second using video.get(cv.CAP_PROP_FPS) : {0}".format(fps))
    
        # Number of frames to capture
        num_frames = 120
    
        print("Capturing {0} frames".format(num_frames))
    
        # Start time
        start = time.time()
    
        # Grab a few frames
        for i in range(0, num_frames) :
            ret, frame = video.read()
    
        # End time
        end = time.time()
    
        # Time elapsed
        seconds = end - start
        print ("Time taken : {0} seconds".format(seconds))
    
        # Calculate frames per second
        fps  = num_frames / seconds
        print("Estimated frames per second : {0}".format(fps))
    
        # Release video
        video.release()

    # =================== INTERFACE ===================
    def analysisVisualization(self, camera1, camera2):
        # Input video streams
        cap0 = cv.VideoCapture(camera1.id, cv.CAP_DSHOW) # DirectShow API Windows
        cap1 = cv.VideoCapture(camera2.id, cv.CAP_DSHOW)
        caps = [cap0, cap1]

        # Set camera resolutions
        for cap in caps:
            cap.set(3, self.settings.frame_width)
            cap.set(4, self.settings.frame_height)

        # Create folder for video_results
        if not os.path.exists('video_results'):
            os.mkdir('video_results')

        trial = input('Write trial number: ')

        # Create VideoWriter objects
        output_name0 = os.path.join('video_results', f'{trial}_L.mp4')
        result0 = cv.VideoWriter(output_name0,
                                cv.VideoWriter_fourcc(*'MP4V'),
                                30, (self.settings.frame_width, self.settings.frame_height))
        
        output_name1 = os.path.join('video_results', f'{trial}_R.mp4')
        result1 = cv.VideoWriter(output_name1,
                                cv.VideoWriter_fourcc(*'MP4V'),
                                30, (self.settings.frame_width, self.settings.frame_height))  

        # Create body keypoint detector object
        pose = mp_pose.Pose(model_complexity=2, min_detection_confidence=0.7, min_tracking_confidence=0.5)

        # Keep track of frame number
        frame_num = 0
        angle_array = [0, 0, 0, 0, 0, 0]

        while True:

            # Keep track of current frame
            frame_num += 1

            # Read frames from stream
            ret0, frame0 = cap0.read()
            ret1, frame1 = cap1.read()

            if not ret0 or not ret1:
                print('Error reading video')
                break            

            # Turn BGR image to RGB
            frame = cv.cvtColor(frame1, cv.COLOR_BGR2RGB)
            # Mark image as not writeable to improve performance
            frame.flags.writeable = False
            # Process results
            results = pose.process(frame)
            # Reverse changes
            frame.flags.writeable = True
            frame = cv.cvtColor(frame, cv.COLOR_RGB2BGR)

            # Display MPP results
            mp_drawing.draw_landmarks(frame, results.pose_landmarks, mp_pose.POSE_CONNECTIONS,
                                      landmark_drawing_spec=mp_drawing_style.get_default_pose_landmarks_style())

            frame_small = cv.resize(frame, (960, 540))
            cv.imshow('MediaPipe results', frame_small)

            # Check for keypoints detection
            pose_array = []
            if results.pose_landmarks:
                print('_' * 25)
                print('Results for frame: ', frame_num)
                for kpt in self.pose_keypoints:

                    x = float(results.pose_landmarks.landmark[self.pose_keypoints[kpt]].x)
                    y = 1 - float(results.pose_landmarks.landmark[self.pose_keypoints[kpt]].y)
                    #z = float(results.pose_landmarks.landmark[self.pose_keypoints[kpt]].z)
                    z = 0
                    xyz = np.array([x, y, z])

                    pose_array.append(np.array([x, y, z], dtype=float))
                    self.pose_dict[kpt].append([x, y, z])

            else:  
                pose_array = np.array([[-1, -1, -1]] * len(self.pose_keypoints))
                for kpt in self.pose_keypoints:
                    self.pose_dict[kpt].append([-1, -1, -1])

            # ROS2: Publish message
            self.publisher.publish_skeleton(pose_array)


            # ===================  SAVE VIDEOS  ======================
            # Save videos
            result0.write(frame0)
            result1.write(frame1)

            # Show video results
            comb = np.concatenate((frame0, frame1), axis=1)
            comb = cv.resize(comb, (1280, 360))
            cv.imshow('Video results', comb)

            k = cv.waitKey(1)
            if k & 0xFF == 27: break  # 27 is ESC key


        cv.destroyAllWindows()
        self.publisher.destroy_node()
        for cap in caps:
            cap.release()
        result0.release()
        result1.release()

        print("Videos saved successfully!\n")  
   

class Gait:
    def __init__(self):

        # KINEMATICS

        self.r_knee_angle = []
        self.r_ankle_angle = []
        self.r_hip_angle = []

        self.l_knee_angle = []
        self.l_ankle_angle = []
        self.l_hip_angle = []

        # SPATIOTEMPORAL

        self.stride_length = []
        self.stride_time = []
        self.swing_time = []
    
    def getAngle2D(self, joint, width, height, side, ankle=True):
        pt1 = (joint[0].x * width, joint[0].y * height, 0) 
        pt2 = (joint[1].x * width, joint[1].y * height, 0)
        pt3 = (joint[2].x * width, joint[2].y * height, 0)

        a = np.array(pt3) - np.array(pt2)
        b = np.array(pt1) - np.array(pt2)

        # FOR ANKLE ANGLE -> TO BE DONE
        if ankle:
            if side == 'R': # when analyzing RIGHT side
                b = np.array([-b[1], b[0], 0]) # perpendicular vector clockwise
            else:
                b = np.array([b[1], -b[0], 0]) # perpendicular vector counterclockwise
                cross_prod = np.cross(b, a) # cross product
        else:
            # FOR HIP ANGLE
            if side == 'R':
                cross_prod = np.cross(b, a)
            else:
                cross_prod = np.cross(a, b)

        # Reference vector in z (+ towards the camera)
        ref = np.array([0, 0, 1])

        if np.dot(cross_prod, ref) < 0:
            sign = -1
        else:
            sign = 1

        # Calculate the dot product of the two vectors
        dot_prod = np.dot(a, b)

        # Calculate the magnitudes of the two vectors
        mag_vec1 = np.linalg.norm(a)
        mag_vec2 = np.linalg.norm(b)

        # Calculate the angle between the two vectors in radians
        angle_rad = np.arccos(dot_prod / (mag_vec1 * mag_vec2))

        # Convert the angle to degrees
        angle_deg = angle_rad * 180 / np.pi

        return angle_deg, sign

    def getAngle3D(self, joint):

        pt1 = (joint[0][0] , joint[0][1], joint[0][1]) 
        pt2 = (joint[1][0] , joint[1][1], joint[1][1]) 
        pt3 = (joint[2][0] , joint[2][1], joint[2][1]) 

        a = np.array(pt3) - np.array(pt2)
        b = np.array(pt1) - np.array(pt2)


        # Calculate the dot product of the two vectors
        dot_prod = np.dot(a, b)

        # Calculate the magnitudes of the two vectors
        mag_vec1 = np.linalg.norm(a)
        mag_vec2 = np.linalg.norm(b)


        # Calculate the angle between the two vectors in radians
        angle_rad = np.arccos(dot_prod / (mag_vec1 * mag_vec2))

        # Convert the angle to degrees
        angle_deg = angle_rad * 180 / np.pi

        print('Angle: ', angle_deg)
        print('_____________________________________\n')
        return angle_deg

    def init_AnglePlot(self):
        
        # Initialize the figure and axis for the plot
        plt.ion()
        fig, (ax1, ax2, ax3) = plt.subplots(3, 1)
        ax1.grid(True)
        ax2.grid(True)
        ax3.grid(True)

        # Define axis limits
        ax1.set_xlim(0, 30000)
        ax1.set_ylim(0, 80)
        ax2.set_xlim(0, 30000)
        ax2.set_ylim(0, 40)
        ax3.set_xlim(0, 30000)
        ax3.set_ylim(0, 180)

        ax1.title.set_text('Knee Flexion')
        ax2.title.set_text('Hip Flexion')
        ax3.title.set_text('Ankle Angle')

        # Initialize the line object for the plot
        line1, = ax1.plot([], [])
        line2, = ax2.plot([], [])
        line3, = ax3.plot([], [])


        # Setting x-axis label and y-axis label
        fig.supxlabel("time (ms)")
        fig.supylabel("angle (º)")

        return line1, line2, line3, ax1, ax2, ax3

    def PoseEstimation2D(self, path):
        
        # Load video
        video = cv.VideoCapture(path)
        side = input('Which side of the body you want to analyze? (R/L): ')

        # Get duration of video
        frame_count = video.get(cv.CAP_PROP_FRAME_COUNT)
        fps = 30
        video_duration = frame_count/fps

        # Get height and width
        w = int(video.get(cv.CAP_PROP_FRAME_WIDTH))
        h = int(video.get(cv.CAP_PROP_FRAME_HEIGHT))

        # Get rescaled height and width for visualization purposes
        scale_percent = 30  # percent of original size
        width = int(w * scale_percent / 100)
        height = int(h * scale_percent / 100)

        # Build plot for joint angles
        knee_line, hip_line, ankle_line, ax1, ax2, ax3 = self.init_AnglePlot()
        
        # Create an empty list to store the values for the angles
        knee_angle = []
        hip_angle = []
        ankle_angle = []

        filt_knee_angle = []
        filt_hip_angle = []
        filt_ankle_angle = []

        # Define extra parameters
        frame_num = -1            # set counter to 0
        time = []                # start time vector
        
        # Run MediaPipe Pose
        with mp_pose.Pose(static_image_mode=False, model_complexity=2, min_detection_confidence=0.75, min_tracking_confidence=0.5) as pose:
            while True:
                # Keep trak of current frame
                frame_num += 1
                # Turn to time (ms) 
                new_time = (frame_num/fps)*1000
                time.append(new_time)

                # Read video
                ret, frame = video.read()
                if not ret:
                    print('Error reading video')
                    break

                # Process image with MediaPipe Pose model
                frame = cv.cvtColor(frame, cv.COLOR_BGR2RGB)
                frame.flags.writeable = False

                results = pose.process(frame)
                
                frame.flags.writeable = True
                frame = cv.cvtColor(frame, cv.COLOR_RGB2BGR)

                # Extract keypoints of interest (normalized)
                if results.pose_landmarks:
                    if side == 'R':
                        knee = [results.pose_landmarks.landmark[i] for i in [24, 26, 28]]
                        hip = [results.pose_landmarks.landmark[i] for i in [12, 24, 26]]
                        ankle = [results.pose_landmarks.landmark[i] for i in [26, 28, 32]]
                    elif side == 'L':
                        knee = [results.pose_landmarks.landmark[i] for i in [23, 25, 27]]
                        hip = [results.pose_landmarks.landmark[i] for i in [11, 23, 25]]
                        ankle = [results.pose_landmarks.landmark[i] for i in [25, 27, 31]]
                    
                
                    # Generate a new value for the vector
                    knee_new, _  = 180 - np.array(self.getAngle2D(knee, w, h, side, ankle=False))
                    knee_angle.append(knee_new)

                    hip_new, sign = self.getAngle2D(hip, w, h, side, ankle=False)
                    hip_new = 180 - np.array(hip_new)
                    hip_new = hip_new * sign
                    hip_angle.append(hip_new)

                    ankle_new, _ = np.array(self.getAngle2D(ankle, w, h, side, ankle=False))
                    ankle_angle.append(ankle_new)
    
                
                else: 
                    knee_angle.append(0)
                    hip_angle.append(0)
                    ankle_angle.append(0)
                
                # Draw results
                mp_drawing.draw_landmarks(frame, results.pose_landmarks, mp_pose.POSE_CONNECTIONS,
                                        mp_drawing.DrawingSpec(color=(128, 0, 250), thickness=2, circle_radius=3),
                                        mp_drawing.DrawingSpec(color=(255, 255, 255), thickness=2))

                cv.namedWindow('MediaPipe results', cv.WINDOW_NORMAL & cv.WINDOW_KEEPRATIO)
                cv.resizeWindow('MediaPipe results', width, height)
                cv.imshow('MediaPipe results', frame)

                k = cv.waitKey(1)

                if k == 27:  # Press ESC to quit
                    break

                # Apply Low Pass filter and plot results
                if len(knee_angle) > 15:

                    nyquist_freq = 0.5 * fps
                    cutoff_freq = 5.0
                    b, a = butter(4, cutoff_freq/nyquist_freq, 'low')
                    filt_knee_angle = filtfilt(b, a, knee_angle)
                    filt_hip_angle = filtfilt(b, a, hip_angle)
                    filt_ankle_angle = filtfilt(b, a, ankle_angle)

                    plt.figure(1)
                    knee_line.set_data(time, filt_knee_angle)
                    ax1.figure.canvas.draw()

                    hip_line.set_data(time, filt_hip_angle)
                    ax2.figure.canvas.draw()

                    ankle_line.set_data(time, filt_ankle_angle)
                    ax3.figure.canvas.draw()

        cv.destroyAllWindows()
        print('Process finished')
        return time, filt_knee_angle, filt_hip_angle, filt_ankle_angle, knee_angle, hip_angle, ankle_angle

    def PoseEstimation3D(self, skeleton, path0, path1, camera1, camera2):
        # Load video
        video0 = cv.VideoCapture(path0)
        video1 = cv.VideoCapture(path1)

        # Get height and width
        w0 = int(video0.get(cv.CAP_PROP_FRAME_WIDTH))
        h0 = int(video0.get(cv.CAP_PROP_FRAME_HEIGHT))
        w1 = int(video1.get(cv.CAP_PROP_FRAME_WIDTH))
        h1 = int(video1.get(cv.CAP_PROP_FRAME_HEIGHT))

        # Create body keypoint detector objects
        pose0 = mp_pose.Pose(model_complexity=2, min_detection_confidence=0.7, min_tracking_confidence=0.5)
        pose1 = mp_pose.Pose(model_complexity=2, min_detection_confidence=0.7, min_tracking_confidence=0.5)
        
        # Create an empty list to store the values for the angles
        knee_angle = []
        hip_angle = []
        ankle_angle = []

        filt_knee_angle = []
        filt_hip_angle = []
        filt_ankle_angle = []

        # Define extra parameters
        frame_num = -1            # set counter
        time = []                 # start time vector
        fps = 30    

        while True:
            frame_num += 1
            new_time = (frame_num/fps)*1000
            time.append(new_time)

            # Read frames from stream
            ret0, frame0 = video0.read()
            ret1, frame1 = video1.read()

            if not ret0 or not ret1:
                print('Error reading video')
                break

            # Turn BGR image to RGB
            frame0 = cv.cvtColor(frame0, cv.COLOR_BGR2RGB)
            frame1 = cv.cvtColor(frame1, cv.COLOR_BGR2RGB)

            # Mark images as not writeable
            frame0.flags.writeable = False
            frame1.flags.writeable = False

            # Process results
            results0 = pose0.process(frame0)
            results1 = pose1.process(frame1)

            # Reverse changes
            frame0.flags.writeable = True
            frame1.flags.writeable = True
            frame0 = cv.cvtColor(frame0, cv.COLOR_RGB2BGR)
            frame1 = cv.cvtColor(frame1, cv.COLOR_RGB2BGR)

            # Display results
            mp_drawing.draw_landmarks(frame0, results0.pose_landmarks, mp_pose.POSE_CONNECTIONS,
                                    landmark_drawing_spec=mp_drawing_style.get_default_pose_landmarks_style())
            mp_drawing.draw_landmarks(frame1, results1.pose_landmarks, mp_pose.POSE_CONNECTIONS,
                                    landmark_drawing_spec=mp_drawing_style.get_default_pose_landmarks_style())

                        
            comb = np.concatenate((frame0, frame1), axis=0)
            comb = cv.resize(comb, (699, 952))
            cv.imshow('MediaPipe results', comb)

            # Check for keypoint detection
            frame0_array = []
            frame1_array = []
            if results0.pose_landmarks and results1.pose_landmarks:
                for kpt in skeleton.pose_keypoints:

                    x0 = int(results0.pose_landmarks.landmark[skeleton.pose_keypoints[kpt]].x * w0)
                    y0 = int(results0.pose_landmarks.landmark[skeleton.pose_keypoints[kpt]].y * h0)
                    frame0_array.append(np.array([x0, y0], dtype=float))

                    camera1.mp_kpts[kpt].append([x0, y0])
                    
                    
                    x1 = int(results1.pose_landmarks.landmark[skeleton.pose_keypoints[kpt]].x * w1)
                    y1 = int(results1.pose_landmarks.landmark[skeleton.pose_keypoints[kpt]].y * h1)
                    frame1_array.append(np.array([x1, y1], dtype=float))

                    camera2.mp_kpts[kpt].append([x1, y1])
                    
            else:
                # Add [-1,-1] to dictionaries if no keypoints are found
                frame0_array = np.array([[-1, -1]] * len(skeleton.pose_keypoints))
                frame1_array = np.array([[-1, -1]] * len(skeleton.pose_keypoints))

            # Transform 2D points to 3D points
            pose_array = []
            for kpt, uv1, uv2 in zip(skeleton.pose_keypoints, frame0_array, frame1_array):
                p3d = np.array([-1, -1, -1])
                if uv1[0] != -1 and uv2[0] != -1:
                    p4d = cv.triangulatePoints(camera1.P, camera2.P, uv1, uv2)  # calculate 3d position of keypoints
                    p3d = (p4d[:3, 0] / p4d[3, 0]).T

                skeleton.pose_dict[kpt].append([p3d[0], p3d[1], p3d[2]])
                pose_array.append(np.array([p3d[0], p3d[2], 1080-p3d[1]], dtype=float))  # Y and Z are exchanged


            # Extract keypoints of interest (normalized)
            left_knee = [pose_array[i] for i in [13, 15, 17]]
            left_hip = [pose_array[i] for i in [1, 13, 15]]
            left_ankle = [pose_array[i] for i in [15, 17, 21]]


            # Generate a new value for the vector
            knee_new = 180 - np.array(self.getAngle3D(left_knee))
            knee_angle.append(knee_new)

            hip_new = 180 - np.array(self.getAngle3D(left_hip))
            hip_angle.append(hip_new)

            ankle_new = - (np.array(self.getAngle3D(left_ankle)) - 110)
            ankle_angle.append(ankle_new)


            # Apply Low Pass filter and plot results
            if len(knee_angle) > 15:

                nyquist_freq = 0.5 * fps
                cutoff_freq = 5.0
                b, a = butter(4, cutoff_freq/nyquist_freq, 'low')
                filt_knee_angle = filtfilt(b, a, knee_angle)
                filt_hip_angle = filtfilt(b, a, hip_angle)
                filt_ankle_angle = filtfilt(b, a, ankle_angle)

        print('Process finished')
        return time, filt_knee_angle, filt_hip_angle, filt_ankle_angle, knee_angle, hip_angle, ankle_angle

    def saveAngularData(self, time, knee_angle, hip_angle, ankle_angle):
        # Define parameters for excel
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        worksheet['A1'] = 'time'
        worksheet['B1'] = 'knee angle'
        worksheet['C1'] = 'hip angle'
        worksheet['D1'] = 'ankle angle'

        count = 0
        for row in range(2, len(knee_angle)+2):
            worksheet.cell(row=row, column=1, value=time[count])
            worksheet.cell(row=row, column=2, value=knee_angle[count])
            worksheet.cell(row=row, column=3, value=hip_angle[count])
            worksheet.cell(row=row, column=4, value=ankle_angle[count])
            count += 1

        # Save workbook
        if not os.path.exists('angular_data'):
            os.mkdir('angular_data')
        
        trial = input("Write trial number: ")
        savename = os.path.join('angular_data', str(trial) + '_kinematics_mp' + '.xlsx')
        workbook.save(savename)  

        print('Data saved to workbook!')

    # ==================== ANALYSES =======================
    def KinematicAnalysis2D(self):
        path = input("Please, introduce the path to the video: ")
        time, filt_knee_angle, filt_hip_angle, filt_ankle_angle, knee_angle, hip_angle, ankle_angle = self.PoseEstimation2D(path)
        self.saveAngularData(time, filt_knee_angle, filt_hip_angle, filt_ankle_angle)
    
    def KinematicAnalysis3D(self, skeleton, camera1, camera2):
        path0 = input("Please, introduce the path to the first video: ")
        path1 = input("Please, introduce the path to the second video: ")
        time, filt_knee_angle, filt_hip_angle, filt_ankle_angle, knee_angle, hip_angle, ankle_angle = self.PoseEstimation3D(skeleton, path0, path1, camera1, camera2)
        self.saveAngularData(time, filt_knee_angle, filt_hip_angle, filt_ankle_angle)

def main(args=None):
    rclpy.init(args=args)
    
    global settings
    settings = Config("C:/dev/ros2_ws/src/pose_estimation/pose_estimation/settings.yaml")

    camera1 = Camera(settings.cam1_id, 'camera0')
    camera2 = Camera(settings.cam2_id, 'camera1')

    skeleton = Skeleton()

    while True:
        # Mostrar el menú de opciones
        print("\nSelect the program mode:")
        print("\n__________        Calibration         __________\n")
        print("1. Calibration")
        print("\n__________   MediaPipe Pose Analysis  __________\n")
        print("2. MediaPipe Pose Video Analysis")
        print("3. MediaPipe Pose Stereo Video Analysis")
        print("4. MediaPipe Pose Single Camera Analysis")
        print("5. MediaPipe Pose Stereo Analysis")
        print("\n__________         Recording           __________\n")
        print("6. Record stereo-cameras")
        print("\n__________       Gait analysis         __________\n")
        print("7. Kinematic Analysis")
        print("\n__________       Visualization         __________\n")
        print("8. Record stereo-cameras + MPP analysis with camera1")

        print("\n0. Exit\n")

        option = input("Your selection: ")

        if option == "1":
            skeleton.calibration(camera1, camera2)

        elif option == "2":
            path = input("Please, introduce the path to your video: ")
            skeleton.analysisVideo(path)

        elif option == "3":
            path0 = input("Please, introduce the path to the video recorded with camera0: ")
            path1 = input("Please, introduce the path to the video recorded with camera1: ")
            camera1.load_calibration()
            camera2.load_calibration()
            skeleton.analysisStereoVideo(path0, path1, camera1, camera2)

        elif option == "4":
            path = input("Please, select the camera you want to use (1/2): ")
            if path == 1:
                skeleton.analysis1Camera(camera1)
            elif path == 2:
                skeleton.analysis1Camera(camera2)
            else:
                print('Introduce a valid option (1/2)')

        elif option == "5":
            camera1.load_calibration()
            camera2.load_calibration()
            skeleton.analysis2Cameras(camera1, camera2)
        
        elif option == "6":
            skeleton.StereoRecord(camera1, camera2)

        elif option == "7":
            print("\n7. Kinematic analysis:")
            print("\t1. Load pre-recorded video (2D kinematic analysis)")
            print("\t2. Load pre-recorded videos (3D kinematic analysis)\n")
            live = input("Your selection: ")
            
            if live == "1":
                gait = Gait()
                gait.KinematicAnalysis2D()
                       
            elif live == "2":
                gait = Gait()
                camera1.load_calibration()
                camera2.load_calibration()
                gait.KinematicAnalysis3D(skeleton, camera1, camera2)

        elif option == "8":
            camera1.load_calibration()
            camera2.load_calibration()
            skeleton.analysisVisualization(camera1, camera2)

        elif option == "0":
            break
        else:
            print("Invalid selection. Plase, introduce a valid command.")

    rclpy.shutdown()


if __name__ == '__main__':
    main()